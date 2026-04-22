from __future__ import annotations

import datetime as dt
from io import BytesIO

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import (
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

from openpyxl import Workbook

from handlers.common import current_month, current_week, get_db, parse_date, require_role
from handlers.menu import send_main_menu


R_PERIOD, R_CUSTOM, R_DIM = 1, 2, 3


def _period_kb(prefix: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("Эта неделя", callback_data=f"{prefix}:week")],
            [InlineKeyboardButton("Этот месяц", callback_data=f"{prefix}:month")],
            [InlineKeyboardButton("Произвольный", callback_data=f"{prefix}:custom")],
            [InlineKeyboardButton("Отмена", callback_data=f"{prefix}:cancel")],
        ]
    )


def _dim_kb(prefix: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("По проектам", callback_data=f"{prefix}:project")],
            [InlineKeyboardButton("По сотрудникам", callback_data=f"{prefix}:user")],
            [InlineKeyboardButton("По клиентам", callback_data=f"{prefix}:client")],
            [InlineKeyboardButton("Отмена", callback_data=f"{prefix}:cancel")],
        ]
    )


def _money_allowed(role: str) -> bool:
    return role in {"admin", "observer"}


@require_role("observer")
async def report_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.effective_message.reply_text("Выбери период:", reply_markup=_period_kb("repP"))
    return R_PERIOD


async def report_period(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    assert q is not None
    await q.answer()
    data = q.data or ""
    if data == "repP:cancel":
        await q.edit_message_text("Ок.")
        await send_main_menu(update, context)
        return ConversationHandler.END

    today = dt.date.today()
    if data == "repP:week":
        p = current_week(today)
        context.user_data["rep_period"] = (p.start, p.end)
        await q.edit_message_text(f"Период: {p.start.isoformat()} .. {p.end.isoformat()}\nВыбери разрез:", reply_markup=_dim_kb("repD"))
        return R_DIM
    if data == "repP:month":
        p = current_month(today)
        context.user_data["rep_period"] = (p.start, p.end)
        await q.edit_message_text(f"Период: {p.start.isoformat()} .. {p.end.isoformat()}\nВыбери разрез:", reply_markup=_dim_kb("repD"))
        return R_DIM
    if data == "repP:custom":
        await q.edit_message_text("Введи период в формате: YYYY-MM-DD YYYY-MM-DD")
        return R_CUSTOM
    return R_PERIOD


async def report_custom(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    assert update.message is not None
    parts = (update.message.text or "").strip().split()
    if len(parts) != 2:
        await update.message.reply_text("Нужно 2 даты: YYYY-MM-DD YYYY-MM-DD")
        return R_CUSTOM
    start = parse_date(parts[0])
    end = parse_date(parts[1])
    if not start or not end or start > end:
        await update.message.reply_text("Некорректный период. Пример: 2026-04-01 2026-04-30")
        return R_CUSTOM
    context.user_data["rep_period"] = (start, end)
    await update.message.reply_text(f"Период: {start.isoformat()} .. {end.isoformat()}\nВыбери разрез:", reply_markup=_dim_kb("repD"))
    return R_DIM


async def report_dim(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    assert q is not None
    await q.answer()
    data = q.data or ""
    if data == "repD:cancel":
        await q.edit_message_text("Ок.")
        await send_main_menu(update, context)
        return ConversationHandler.END
    group_by = data.split(":")[-1]
    start, end = context.user_data["rep_period"]
    db = await get_db(context)
    me = context.user_data["me"]
    rows = await db.report_grouped(start=start, end=end, group_by=group_by, restrict_user_id=None)
    show_money = _money_allowed(me.role)
    abs_map: dict[int, dict[str, int]] = {}
    if group_by == "user":
        for r in await db.absence_counts(start=start, end=end):
            uid = int(r["user_id"])
            reason = str(r["reason"])
            days = int(r["days"] or 0)
            abs_map.setdefault(uid, {})[reason] = days
    lines = [f"Отчёт {start.isoformat()} .. {end.isoformat()} ({group_by})"]
    if not rows:
        lines.append("Данных нет.")
        await q.edit_message_text("\n".join(lines))
        await send_main_menu(update, context)
        return ConversationHandler.END

    for r in rows:
        h = float(r["hours"] or 0)
        suffix = ""
        if group_by == "user":
            uid = int(r["group_id"])
            reasons = abs_map.get(uid, {})
            parts_abs: list[str] = []
            if reasons.get("vacation"):
                parts_abs.append(f"{reasons['vacation']} дн отпуск")
            if reasons.get("sick"):
                parts_abs.append(f"{reasons['sick']} дн болел")
            if reasons.get("dayoff"):
                parts_abs.append(f"{reasons['dayoff']} дн day off")
            if parts_abs:
                suffix = " (" + ", ".join(parts_abs) + ")"
        if show_money:
            ic = float(r["internal_cost"] or 0)
            ec = float(r["external_cost"] or 0)
            lines.append(f"- {r['label']}{suffix}: {h:.2f} ч | себест. {ic:.2f} | клиент. {ec:.2f}")
        else:
            lines.append(f"- {r['label']}{suffix}: {h:.2f} ч")
    await q.edit_message_text("\n".join(lines))
    await send_main_menu(update, context)
    return ConversationHandler.END


report_conversation = ConversationHandler(
    entry_points=[CommandHandler("report", report_entry), CallbackQueryHandler(report_entry, pattern=r"^menu:report$")],
    states={
        R_PERIOD: [CallbackQueryHandler(report_period, pattern=r"^repP:")],
        R_CUSTOM: [MessageHandler(filters.TEXT & ~filters.COMMAND, report_custom)],
        R_DIM: [CallbackQueryHandler(report_dim, pattern=r"^repD:")],
    },
    fallbacks=[],
    name="report",
    persistent=False,
)


@require_role("observer")
async def workload(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    today = dt.date.today()
    p = current_month(today)
    start, end = p.start, p.end
    db = await get_db(context)
    rows = await db.fetchall(
        """
        SELECT u.name AS user_name,
               SUM(t.hours) AS hours
        FROM timelog t
        JOIN users u ON u.id = t.user_id
        WHERE t.date BETWEEN ? AND ?
        GROUP BY u.id
        ORDER BY u.name;
        """,
        (start.isoformat(), end.isoformat()),
    )
    if not rows:
        await update.effective_message.reply_text("Данных нет.")
        return

    lines = [f"Загрузка за {start.isoformat()} .. {end.isoformat()}:"]
    for r in rows:
        user_name = r["user_name"]
        hours = float(r["hours"] or 0)
        lines.append(f"- {user_name}: {hours:.2f} ч")
        prows = await db.fetchall(
            """
            SELECT c.name AS client_name, p.name AS project_name, SUM(t.hours) AS hours
            FROM timelog t
            JOIN projects p ON p.id = t.project_id
            JOIN clients c ON c.id = p.client_id
            JOIN users u ON u.id = t.user_id
            WHERE u.name = ? AND t.date BETWEEN ? AND ?
            GROUP BY p.id
            ORDER BY c.name, p.name;
            """,
            (user_name, start.isoformat(), end.isoformat()),
        )
        for pr in prows:
            lines.append(f"  - {pr['client_name']} / {pr['project_name']}: {float(pr['hours'] or 0):.2f} ч")
    await update.effective_message.reply_text("\n".join(lines))


workload_handler = CommandHandler("workload", workload)


# ---------- admin export ----------


E_PERIOD, E_CUSTOM, E_TYPE, E_USER = 1, 2, 3, 4


def _prev_month(today: dt.date) -> tuple[dt.date, dt.date]:
    first_this = dt.date(today.year, today.month, 1)
    last_prev = first_this - dt.timedelta(days=1)
    first_prev = dt.date(last_prev.year, last_prev.month, 1)
    return first_prev, last_prev


def _export_period_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("Эта неделя", callback_data="expP:week")],
            [InlineKeyboardButton("Этот месяц", callback_data="expP:month")],
            [InlineKeyboardButton("Прошлый месяц", callback_data="expP:prevmonth")],
            [InlineKeyboardButton("Произвольный", callback_data="expP:custom")],
            [InlineKeyboardButton("Отмена", callback_data="expP:cancel")],
        ]
    )


def _export_type_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("По всем сотрудникам", callback_data="expT:employees")],
            [InlineKeyboardButton("По клиентам", callback_data="expT:clients")],
            [InlineKeyboardButton("По отдельному сотруднику", callback_data="expT:user")],
            [InlineKeyboardButton("Отмена", callback_data="expT:cancel")],
        ]
    )


def _export_users_kb(users: list[dict]) -> InlineKeyboardMarkup:
    kb: list[list[InlineKeyboardButton]] = [
        [InlineKeyboardButton(f"{u['name']} ({u['role']})", callback_data=f"expU:{u['id']}")] for u in users
    ]
    kb.append([InlineKeyboardButton("⬅️ Назад", callback_data="expU:back")])
    kb.append([InlineKeyboardButton("Отмена", callback_data="expU:cancel")])
    return InlineKeyboardMarkup(kb)


@require_role("admin")
async def export_entry(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.effective_message.reply_text("Выбери период для Excel:", reply_markup=_export_period_kb())
    return E_PERIOD


async def export_period(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    assert q is not None
    await q.answer()
    data = q.data or ""
    if data == "expP:cancel":
        await q.edit_message_text("Ок.")
        await send_main_menu(update, context)
        return ConversationHandler.END
    today = dt.date.today()
    if data == "expP:week":
        p = current_week(today)
        context.user_data["exp_period"] = (p.start, p.end)
        await q.edit_message_text(
            f"Период: {p.start.isoformat()} .. {p.end.isoformat()}\nВыбери тип выгрузки:",
            reply_markup=_export_type_kb(),
        )
        return E_TYPE
    if data == "expP:month":
        p = current_month(today)
        context.user_data["exp_period"] = (p.start, p.end)
        await q.edit_message_text(
            f"Период: {p.start.isoformat()} .. {p.end.isoformat()}\nВыбери тип выгрузки:",
            reply_markup=_export_type_kb(),
        )
        return E_TYPE
    if data == "expP:prevmonth":
        start, end = _prev_month(today)
        context.user_data["exp_period"] = (start, end)
        await q.edit_message_text(
            f"Период: {start.isoformat()} .. {end.isoformat()}\nВыбери тип выгрузки:",
            reply_markup=_export_type_kb(),
        )
        return E_TYPE
    if data == "expP:custom":
        await q.edit_message_text("Введи период в формате: YYYY-MM-DD YYYY-MM-DD")
        return E_CUSTOM
    return E_PERIOD


async def export_custom(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    assert update.message is not None
    parts = (update.message.text or "").strip().split()
    if len(parts) != 2:
        await update.message.reply_text("Нужно 2 даты: YYYY-MM-DD YYYY-MM-DD")
        return E_CUSTOM
    start = parse_date(parts[0])
    end = parse_date(parts[1])
    if not start or not end or start > end:
        await update.message.reply_text("Некорректный период.")
        return E_CUSTOM
    context.user_data["exp_period"] = (start, end)
    await update.message.reply_text(
        f"Период: {start.isoformat()} .. {end.isoformat()}\nВыбери тип выгрузки:",
        reply_markup=_export_type_kb(),
    )
    return E_TYPE


async def export_type(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    assert q is not None
    await q.answer()
    data = q.data or ""
    if data == "expT:cancel":
        await q.edit_message_text("Ок.")
        await send_main_menu(update, context)
        return ConversationHandler.END

    context.user_data["exp_type"] = data.split(":")[-1]
    if context.user_data["exp_type"] == "user":
        db = await get_db(context)
        users = [dict(r) for r in await db.list_users()]
        await q.edit_message_text("Выбери сотрудника:", reply_markup=_export_users_kb(users))
        return E_USER

    await q.edit_message_text("Готовлю Excel…")
    return await export_build_and_send(update, context)


async def export_user_select(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    q = update.callback_query
    assert q is not None
    await q.answer()
    data = q.data or ""
    if data == "expU:cancel":
        await q.edit_message_text("Ок.")
        await send_main_menu(update, context)
        return ConversationHandler.END
    if data == "expU:back":
        await q.edit_message_text("Выбери тип выгрузки:", reply_markup=_export_type_kb())
        return E_TYPE
    if data.startswith("expU:"):
        user_id = int(data.split(":")[-1])
        context.user_data["exp_user_id"] = user_id
        await q.edit_message_text("Готовлю Excel…")
        return await export_build_and_send(update, context)
    return E_USER


async def export_build_and_send(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    start, end = context.user_data["exp_period"]
    db = await get_db(context)

    kind = str(context.user_data.get("exp_type") or "employees")

    wb = Workbook()
    ws = wb.active

    if kind == "employees":
        ws.title = "Сотрудники"
        ws.append(["Имя", "Роль", "Всего часов", "Себестоимость", "Клиентская стоимость"])
        rows = await db.fetchall(
            """
            SELECT
              u.name AS user_name,
              u.role AS role,
              COALESCE(SUM(t.hours),0) AS hours,
              COALESCE(SUM(t.hours * u.internal_rate),0) AS internal_cost,
              COALESCE(SUM(t.hours * u.external_rate),0) AS external_cost
            FROM users u
            LEFT JOIN timelog t
              ON t.user_id = u.id
             AND t.date BETWEEN ? AND ?
            GROUP BY u.id
            ORDER BY u.name;
            """,
            (start.isoformat(), end.isoformat()),
        )
        for r in rows:
            ws.append(
                [
                    r["user_name"],
                    r["role"],
                    float(r["hours"] or 0),
                    float(r["internal_cost"] or 0),
                    float(r["external_cost"] or 0),
                ]
            )

        det = wb.create_sheet("Детализация")
        det.append(["Сотрудник", "Клиент", "Проект", "Часы", "Себестоимость", "Клиентская стоимость"])
        drows = await db.fetchall(
            """
            SELECT
              u.name AS user_name,
              c.name AS client_name,
              p.name AS project_name,
              SUM(t.hours) AS hours,
              SUM(t.hours * u.internal_rate) AS internal_cost,
              SUM(t.hours * u.external_rate) AS external_cost
            FROM timelog t
            JOIN users u ON u.id = t.user_id
            JOIN projects p ON p.id = t.project_id
            JOIN clients c ON c.id = p.client_id
            WHERE t.date BETWEEN ? AND ?
            GROUP BY u.id, c.id, p.id
            ORDER BY u.name, c.name, p.name;
            """,
            (start.isoformat(), end.isoformat()),
        )
        for r in drows:
            det.append(
                [
                    r["user_name"],
                    r["client_name"],
                    r["project_name"],
                    float(r["hours"] or 0),
                    float(r["internal_cost"] or 0),
                    float(r["external_cost"] or 0),
                ]
            )

    elif kind == "clients":
        ws.title = "Клиенты"
        ws.append(["Клиент", "Всего часов", "Себестоимость", "Клиентская стоимость"])
        rows = await db.fetchall(
            """
            SELECT
              c.name AS client_name,
              SUM(t.hours) AS hours,
              SUM(t.hours * u.internal_rate) AS internal_cost,
              SUM(t.hours * u.external_rate) AS external_cost
            FROM timelog t
            JOIN users u ON u.id = t.user_id
            JOIN projects p ON p.id = t.project_id
            JOIN clients c ON c.id = p.client_id
            WHERE t.date BETWEEN ? AND ?
            GROUP BY c.id
            ORDER BY c.name;
            """,
            (start.isoformat(), end.isoformat()),
        )
        for r in rows:
            ws.append(
                [
                    r["client_name"],
                    float(r["hours"] or 0),
                    float(r["internal_cost"] or 0),
                    float(r["external_cost"] or 0),
                ]
            )

        det = wb.create_sheet("Детализация")
        det.append(["Клиент", "Проект", "Сотрудник", "Часы", "Себестоимость", "Клиентская стоимость"])
        drows = await db.fetchall(
            """
            SELECT
              c.name AS client_name,
              p.name AS project_name,
              u.name AS user_name,
              SUM(t.hours) AS hours,
              SUM(t.hours * u.internal_rate) AS internal_cost,
              SUM(t.hours * u.external_rate) AS external_cost
            FROM timelog t
            JOIN users u ON u.id = t.user_id
            JOIN projects p ON p.id = t.project_id
            JOIN clients c ON c.id = p.client_id
            WHERE t.date BETWEEN ? AND ?
            GROUP BY c.id, p.id, u.id
            ORDER BY c.name, p.name, u.name;
            """,
            (start.isoformat(), end.isoformat()),
        )
        for r in drows:
            det.append(
                [
                    r["client_name"],
                    r["project_name"],
                    r["user_name"],
                    float(r["hours"] or 0),
                    float(r["internal_cost"] or 0),
                    float(r["external_cost"] or 0),
                ]
            )

    elif kind == "user":
        user_id = int(context.user_data["exp_user_id"])
        ws.title = "Отчёт"
        ws.append(["Проект", "Клиент", "Часы", "Себестоимость", "Клиентская стоимость"])
        rows = await db.report_grouped(start=start, end=end, group_by="project", restrict_user_id=user_id)
        for r in rows:
            ws.append(
                [
                    r["label"],
                    r["client_name"],
                    float(r["hours"] or 0),
                    float(r["internal_cost"] or 0),
                    float(r["external_cost"] or 0),
                ]
            )
    else:
        ws.title = "Отчёт"
        ws.append(["Данных нет"])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    if kind == "employees":
        kind_slug = "employees"
    elif kind == "clients":
        kind_slug = "clients"
    elif kind == "user":
        kind_slug = f"user_{int(context.user_data.get('exp_user_id') or 0)}"
    else:
        kind_slug = "export"

    filename = f"report_{kind_slug}_{start.isoformat()}_{end.isoformat()}.xlsx"
    await update.effective_message.reply_document(document=bio, filename=filename)
    await send_main_menu(update, context)
    return ConversationHandler.END


export_conversation = ConversationHandler(
    entry_points=[CommandHandler("export", export_entry), CallbackQueryHandler(export_entry, pattern=r"^menu:export$")],
    states={
        E_PERIOD: [CallbackQueryHandler(export_period, pattern=r"^expP:")],
        E_CUSTOM: [MessageHandler(filters.TEXT & ~filters.COMMAND, export_custom)],
        E_TYPE: [CallbackQueryHandler(export_type, pattern=r"^expT:")],
        E_USER: [CallbackQueryHandler(export_user_select, pattern=r"^expU:")],
    },
    fallbacks=[],
    name="export",
    persistent=False,
)

